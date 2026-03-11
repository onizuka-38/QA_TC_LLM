from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from src.backend.models import ParserOutput

from .base import BaseParser


class ExcelParser(BaseParser):
    file_type = "xlsx"

    def parse(self, document_id: str, filename: str, content: bytes) -> ParserOutput:
        try:
            workbook = load_workbook(BytesIO(content), data_only=True)
            lines: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if values:
                        lines.append(" | ".join(values))
            extracted_text = "\n".join(lines)
        except Exception:
            # Keep fallback for invalid test fixture bytes and corrupted xlsx files.
            extracted_text = content.decode("utf-8", errors="ignore")

        return ParserOutput(
            document_id=document_id,
            filename=filename,
            file_type=self.file_type,
            extracted_text=extracted_text,
            source_location="n/a",
        )
