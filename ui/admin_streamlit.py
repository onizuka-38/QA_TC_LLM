from __future__ import annotations

from io import BytesIO
from typing import Any

import httpx
import streamlit as st
from openpyxl import load_workbook


st.set_page_config(page_title="QA TC Workspace", layout="wide")

st.markdown(
    """
<style>
    .stApp {
        background: linear-gradient(180deg, #f7f9fc 0%, #eef3fb 100%);
    }
    .block-title {
        font-size: 1.1rem;
        font-weight: 700;
        margin-bottom: 0.4rem;
    }
    .muted {
        color: #53627c;
        font-size: 0.92rem;
    }
    .pill {
        display: inline-block;
        border: 1px solid #d9e1f0;
        border-radius: 999px;
        padding: 0.2rem 0.6rem;
        margin: 0.15rem 0.25rem 0.15rem 0;
        background: #ffffff;
        font-size: 0.82rem;
        color: #1b2a4a;
    }
</style>
    """,
    unsafe_allow_html=True,
)


def init_state() -> None:
    defaults: dict[str, object] = {
        "document_ids": [],
        "requirements": [],
        "selected_requirement_ids": [],
        "selected_requirement_ids_widget": [],
        "request_id": "",
        "export_bytes": b"",
        "chat_log": [],
        "draft_cases": [],
        "review_completed": False,
        "jobs_result": "",
        "validation_result": "",
        "rtm_result": "",
        "chat_prompt_input": "",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def render_export_preview(content: bytes) -> None:
    wb = load_workbook(BytesIO(content), data_only=True)
    if "TC" in wb.sheetnames:
        ws_tc = wb["TC"]
        rows_tc = list(ws_tc.iter_rows(values_only=True))
        if rows_tc:
            headers = [str(h) if h is not None else "" for h in rows_tc[0]]
            preview_tc: list[dict[str, Any]] = []
            for row in rows_tc[1:11]:
                preview_tc.append({headers[i]: row[i] for i in range(len(headers))})
            st.markdown("**TC 미리보기**")
            st.dataframe(preview_tc, width="stretch")

    if "RTM" in wb.sheetnames:
        ws_rtm = wb["RTM"]
        rows_rtm = list(ws_rtm.iter_rows(values_only=True))
        if rows_rtm:
            headers = [str(h) if h is not None else "" for h in rows_rtm[0]]
            preview_rtm: list[dict[str, Any]] = []
            for row in rows_rtm[1:11]:
                preview_rtm.append({headers[i]: row[i] for i in range(len(headers))})
            st.markdown("**RTM 미리보기**")
            st.dataframe(preview_rtm, width="stretch")


init_state()

st.title("QA/TC Workspace")
st.caption("Document-grounded chat + structured QA workflow")

api_base_url = st.sidebar.text_input("FastAPI Base URL", value="http://127.0.0.1:8010")
requested_by = st.sidebar.text_input("Requested By", value="qa_user")
st.sidebar.caption("Ports: vLLM 8001 / FastAPI 8010 / Streamlit 8510")

status_col1, status_col2, status_col3 = st.columns(3)
status_col1.metric("Documents", len(st.session_state.document_ids))
status_col2.metric("Requirements", len(st.session_state.requirements))
status_col3.metric("Selected REQ", len(st.session_state.selected_requirement_ids))

left_col, right_col = st.columns([1.05, 1.25], gap="large")

with left_col:
    with st.container(border=True):
        st.markdown('<div class="block-title">1) 문서 업로드 + Requirement 선택</div>', unsafe_allow_html=True)
        st.markdown('<div class="muted">문서를 올리고 requirement를 자동 추출한 뒤 선택합니다.</div>', unsafe_allow_html=True)

        uploaded_files = st.file_uploader(
            "문서 첨부 (pdf/docx/xlsx)",
            type=["pdf", "docx", "xlsx"],
            accept_multiple_files=True,
        )
        upload_col, req_col = st.columns(2)
        with upload_col:
            if st.button("업로드", type="primary", width="stretch"):
                if not uploaded_files:
                    st.warning("업로드할 파일을 선택하세요.")
                else:
                    files = [("files", (f.name, f.getvalue(), "application/octet-stream")) for f in uploaded_files]
                    data = {"requested_by": requested_by}
                    with httpx.Client(timeout=90.0) as client:
                        resp = client.post(f"{api_base_url}/documents/upload", files=files, data=data)
                    if resp.status_code == 200:
                        payload = resp.json()
                        docs = payload.get("documents", [])
                        st.session_state.document_ids = [doc["document_id"] for doc in docs]
                        st.session_state.review_completed = False
                        st.success("업로드 성공")
                    else:
                        st.error(f"업로드 실패: {resp.status_code}")
                        st.code(resp.text)
        with req_col:
            if st.button("Requirement 추출", width="stretch"):
                all_requirements: list[dict[str, Any]] = []
                with httpx.Client(timeout=90.0) as client:
                    for document_id in st.session_state.document_ids:
                        resp = client.get(f"{api_base_url}/documents/{document_id}/requirements")
                        if resp.status_code == 200:
                            all_requirements.extend(resp.json().get("requirements", []))
                by_id: dict[str, dict[str, Any]] = {}
                for item in all_requirements:
                    rid = str(item.get("requirement_id", ""))
                    if rid and rid not in by_id:
                        by_id[rid] = item
                st.session_state.requirements = list(by_id.values())
                refreshed = {item["requirement_id"] for item in st.session_state.requirements}
                st.session_state.selected_requirement_ids_widget = [
                    rid for rid in st.session_state.selected_requirement_ids_widget if rid in refreshed
                ]

        if st.session_state.document_ids:
            st.markdown("**문서 ID**")
            doc_pills = "".join([f'<span class="pill">{doc_id}</span>' for doc_id in st.session_state.document_ids])
            st.markdown(doc_pills, unsafe_allow_html=True)

        req_options = [item["requirement_id"] for item in st.session_state.requirements]
        st.multiselect(
            "Requirement 선택",
            options=req_options,
            key="selected_requirement_ids_widget",
            placeholder="Requirement를 선택하세요.",
        )
        st.session_state.selected_requirement_ids = list(st.session_state.selected_requirement_ids_widget)

    with st.container(border=True):
        st.markdown('<div class="block-title">2) 챗봇형 문서 QA</div>', unsafe_allow_html=True)
        st.markdown('<div class="muted">답변은 선택한 requirement와 문서 근거(source_chunks) 기반으로만 생성됩니다.</div>', unsafe_allow_html=True)

        for row in st.session_state.chat_log[-12:]:
            with st.chat_message("user"):
                st.write(str(row.get("question", "")))
            with st.chat_message("assistant"):
                st.write(str(row.get("answer", "")))
                st.caption(f"source_chunks: {row.get('source_chunks', [])}")

        with st.form("chat_query_form", clear_on_submit=True):
            chat_prompt = st.text_area(
                "질문 입력",
                key="chat_prompt_input",
                placeholder="예: REQ-100 기준 빠진 예외 케이스가 뭐야?",
                height=90,
            )
            send_chat = st.form_submit_button("질문 보내기", type="primary", width="stretch")

        if send_chat:
            if not st.session_state.document_ids:
                st.warning("먼저 문서를 업로드하세요.")
            elif not chat_prompt.strip():
                st.warning("질문을 입력하세요.")
            else:
                payload = {
                    "document_ids": st.session_state.document_ids,
                    "selected_requirement_ids": st.session_state.selected_requirement_ids,
                    "user_prompt": chat_prompt.strip(),
                    "requested_by": requested_by,
                }
                try:
                    with httpx.Client(timeout=300.0) as client:
                        resp = client.post(f"{api_base_url}/chat/query", json=payload)
                except httpx.TimeoutException:
                    st.error("chat 실패: timeout (vLLM 응답 지연)")
                else:
                    if resp.status_code == 200:
                        data = resp.json()
                        st.session_state.chat_log.append(
                            {
                                "question": chat_prompt.strip(),
                                "answer": data.get("answer", ""),
                                "source_chunks": data.get("source_chunks", []),
                            }
                        )
                        st.rerun()
                    else:
                        st.error(f"chat 실패: {resp.status_code}")
                        st.code(resp.text)

with right_col:
    with st.container(border=True):
        st.markdown('<div class="block-title">3) 구조화 QA/TC 생성</div>', unsafe_allow_html=True)
        st.markdown('<div class="muted">선택된 requirement 기준으로 TC draft를 생성합니다.</div>', unsafe_allow_html=True)

        user_prompt = st.text_area(
            "생성 보조 입력(user_prompt)",
            value="선택된 requirement_id 각각에 대해 정상/오류/예외 관점으로 3~5개 TC를 제안해줘.",
            height=100,
        )
        target_case_count = st.selectbox("목표 TC 개수", options=[3, 4, 5], index=0)
        if st.button("Generate", type="primary", width="stretch"):
            payload = {
                "document_ids": st.session_state.document_ids,
                "requirement_ids": st.session_state.selected_requirement_ids,
                "user_prompt": user_prompt,
                "target_case_count": target_case_count,
                "requested_by": requested_by,
            }
            try:
                with st.spinner("생성 중... (모델 응답에 따라 1~4분 소요 가능)"):
                    with httpx.Client(timeout=600.0) as client:
                        resp = client.post(f"{api_base_url}/tc/generate", json=payload)
            except httpx.TimeoutException:
                st.error("생성 실패: timeout (서버 처리 시간 초과)")
            else:
                if resp.status_code == 200:
                    result = resp.json()
                    st.session_state.request_id = result.get("request_id", "")
                    st.session_state.review_completed = False
                    st.success("생성 요청 완료")
                    st.json(result)
                    with httpx.Client(timeout=60.0) as client:
                        draft_resp = client.get(f"{api_base_url}/tc/drafts/{st.session_state.request_id}")
                    if draft_resp.status_code == 200:
                        st.session_state.draft_cases = draft_resp.json().get("cases", [])
                else:
                    st.error(f"생성 실패: {resp.status_code}")
                    st.code(resp.text)

        st.session_state.request_id = st.text_input("request_id", value=st.session_state.request_id)

        row1_col1, row1_col2, row1_col3 = st.columns(3)
        with row1_col1:
            if st.button("jobs 조회", width="stretch"):
                with httpx.Client(timeout=60.0) as client:
                    resp = client.get(f"{api_base_url}/jobs/{st.session_state.request_id}")
                st.session_state.jobs_result = resp.text
        with row1_col2:
            if st.button("validation 조회", width="stretch"):
                with httpx.Client(timeout=60.0) as client:
                    resp = client.get(f"{api_base_url}/validation/{st.session_state.request_id}")
                st.session_state.validation_result = resp.text
        with row1_col3:
            if st.button("rtm 조회", width="stretch"):
                with httpx.Client(timeout=60.0) as client:
                    resp = client.get(f"{api_base_url}/rtm/{st.session_state.request_id}")
                st.session_state.rtm_result = resp.text

        with st.expander("조회 결과", expanded=False):
            if st.session_state.jobs_result:
                st.markdown("**jobs**")
                st.code(st.session_state.jobs_result)
            if st.session_state.validation_result:
                st.markdown("**validation**")
                st.code(st.session_state.validation_result)
            if st.session_state.rtm_result:
                st.markdown("**rtm**")
                st.code(st.session_state.rtm_result)

    with st.container(border=True):
        st.markdown('<div class="block-title">4) Draft 편집 + Review</div>', unsafe_allow_html=True)
        action_col1, action_col2 = st.columns(2)
        with action_col1:
            if st.button("draft 조회", width="stretch"):
                with httpx.Client(timeout=60.0) as client:
                    resp = client.get(f"{api_base_url}/tc/drafts/{st.session_state.request_id}")
                if resp.status_code == 200:
                    st.session_state.draft_cases = resp.json().get("cases", [])
                else:
                    st.error(f"draft 조회 실패: {resp.status_code}")
                    st.code(resp.text)
        with action_col2:
            if st.button("검토 완료(확정)", width="stretch"):
                with httpx.Client(timeout=120.0) as client:
                    resp = client.post(
                        f"{api_base_url}/tc/review/{st.session_state.request_id}/complete",
                        json={"requested_by": requested_by},
                    )
                st.session_state.review_completed = resp.status_code == 200
                if resp.status_code == 200:
                    st.success("검토 완료 처리됨")
                else:
                    st.error(f"검토 완료 실패: {resp.status_code}")
                    st.code(resp.text)
                    if resp.status_code == 409:
                        with httpx.Client(timeout=60.0) as client:
                            validation_resp = client.get(f"{api_base_url}/validation/{st.session_state.request_id}")
                        if validation_resp.status_code == 200:
                            st.warning("검증 실패 사유")
                            st.json(validation_resp.json())

        if st.session_state.draft_cases:
            edited = st.data_editor(st.session_state.draft_cases, width="stretch")
            if st.button("draft 저장", width="stretch"):
                payload = {"cases": edited, "requested_by": requested_by}
                with httpx.Client(timeout=120.0) as client:
                    resp = client.put(f"{api_base_url}/tc/drafts/{st.session_state.request_id}", json=payload)
                if resp.status_code == 200:
                    st.session_state.draft_cases = resp.json().get("cases", [])
                    st.success("draft 저장 완료")
                else:
                    st.error(f"draft 저장 실패: {resp.status_code}")
                    st.code(resp.text)
        else:
            st.info("draft가 비어 있습니다. jobs/validation 결과를 먼저 확인해 주세요.")

    with st.container(border=True):
        st.markdown('<div class="block-title">5) Export + Preview</div>', unsafe_allow_html=True)
        if st.button("export 조회", disabled=not st.session_state.review_completed, width="stretch"):
            with httpx.Client(timeout=120.0) as client:
                resp = client.get(f"{api_base_url}/exports/{st.session_state.request_id}")
            content_type = resp.headers.get("content-type", "")
            if resp.status_code == 200 and "spreadsheetml.sheet" in content_type:
                st.session_state.export_bytes = resp.content
                st.success("xlsx export 수신 성공")
            else:
                st.error("export 실패")
                st.code(resp.text)

        if st.session_state.export_bytes:
            st.download_button(
                label="tc_rtm.xlsx 다운로드",
                data=st.session_state.export_bytes,
                file_name="tc_rtm.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
            render_export_preview(st.session_state.export_bytes)
